import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os

def create_excel_file(folder_path, engineer_name, store_name, address, date_str):
    safe_date_str = date_str.replace(':', '_')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Выявленные дефекты"

    # Записываем основную информацию без переноса текста
    ws['A1'] = f"ФИО инженера: {engineer_name}"
    ws['A2'] = f"Название магазина: {store_name}"
    ws['A3'] = f"Адрес магазина: {address}"
    ws['A4'] = f"Дата проведения аудита: {safe_date_str}"

    # Создаем заголовки таблицы
    headers = ["№ дефекта", "Вид дефекта", "Описание выявленного дефекта", "Работы необходимые по устранению выявленного дефекта"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Устанавливаем жесткую ширину для столбцов "Описание выявленного дефекта" и "Работы необходимые по устранению выявленного дефекта"
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25

    # Применяем границы к таблице
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws['A6:D6']:
        for cell in row:
            cell.border = thin_border

    # Устанавливаем автоматическую высоту строк для заголовков
    for row in ws.iter_rows(min_row=6, max_row=6, min_col=1, max_col=ws.max_column):
        max_height = 0
        for cell in row:
            cell_height = len(str(cell.value).split('\n'))
            if cell_height > max_height:
                max_height = cell_height
        ws.row_dimensions[6].height = max_height * 35

    # Сохраняем файл
    file_path = os.path.join(folder_path, f"Выявленные дефекты_{store_name}_{safe_date_str}.xlsx")
    wb.save(file_path)
    return file_path

def append_defect_to_excel(file_path, defect_number, defect_type, description, work_needed):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    row = ws.max_row + 1

    # Устанавливаем значения ячеек
    ws.cell(row=row, column=1, value=defect_number).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(row=row, column=2, value=defect_type).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(row=row, column=3, value=description).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.cell(row=row, column=4, value=work_needed).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Применяем границы к ячейкам
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[row]:
        cell.border = thin_border

    # Устанавливаем автоматическую высоту строк
    max_height = 0
    for cell in ws[row]:
        cell_height = len(str(cell.value).split('\n'))
        if cell_height > max_height:
            max_height = cell_height
    ws.row_dimensions[row].height = max_height * 15

    wb.save(file_path)
